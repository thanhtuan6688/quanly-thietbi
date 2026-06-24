from __future__ import annotations

import hashlib
import hmac
import io
import json
import mimetypes
import os
import platform
import secrets
import socket
import subprocess
import sys
import threading
import time
import warnings
import zipfile
from copy import copy
from datetime import datetime
from email import policy
from email.parser import BytesParser
from functools import lru_cache
from http import HTTPStatus
from http.cookies import SimpleCookie
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, unquote, urlparse
from uuid import uuid4

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - startup guard
    raise SystemExit("Thieu thu vien openpyxl. Hay chay bang runtime Python di kem Codex.") from exc

warnings.filterwarnings(
    "ignore",
    message="Data Validation extension is not supported and will be removed",
    category=UserWarning,
    module="openpyxl.worksheet._reader",
)


APP_ROOT = Path(__file__).resolve().parent
STATIC_DIR = APP_ROOT / "static"
ASSETS_DIR = APP_ROOT / "assets"
DATA_ROOT = Path(os.environ.get("DEVICE_INVENTORY_DATA_ROOT", APP_ROOT)).resolve()
DATA_DIR = Path(os.environ.get("DEVICE_INVENTORY_DATA_DIR", DATA_ROOT / "data")).resolve()
UPLOAD_DIR = Path(os.environ.get("DEVICE_INVENTORY_UPLOAD_DIR", DATA_ROOT / "uploads")).resolve()
EXPORT_DIR = Path(os.environ.get("DEVICE_INVENTORY_EXPORT_DIR", DATA_ROOT / "exports")).resolve()
CLIENT_INFO_DIR = Path(os.environ.get("DEVICE_INVENTORY_CLIENT_INFO_DIR", DATA_ROOT / "client-info")).resolve()
TEMPLATE_PATH = ASSETS_DIR / "Ra_Soat_Thiet_Bi_CNTT.xlsx"
STORE_PATH = DATA_DIR / "inventory.json"
CLIENT_INVENTORY_PATH = DATA_DIR / "client-inventory.json"
MAX_UPLOAD_BYTES = 50 * 1024 * 1024
MAX_CLIENT_INFO_BYTES = 256 * 1024
CLIENT_INFO_TTL_SECONDS = 30 * 60
CLIENT_INVENTORY_TTL_SECONDS = 60 * 60
ADMIN_PASSWORD = os.environ.get("DEVICE_INVENTORY_ADMIN_PASSWORD", "250389")
ADMIN_COOKIE_NAME = "device_inventory_admin"
ADMIN_SESSION_SECONDS = 12 * 60 * 60
ADMIN_SESSION_SECRET = secrets.token_bytes(32)
CLIENT_INVENTORY_LOCK = threading.RLock()

COMMON_REQUIRED = [
    "Tên Tỉnh",
    "Tên Bưu cục",
    "Họ và tên người sử dụng",
    "Bộ phận / Phòng ban",
    "Mã vật tư",
    "Tên tài sản\n(Theo danh mục CCDC)",
    "Tình trạng",
]

CATEGORY_ORDER = [
    "3. Máy Tính",
    "4A. Máy In Laser",
    "4B. Máy In Nhiệt",
    "4C. Máy Photocopy",
    "5. Đầu Đọc Mã Vạch",
    "6. Thiết Bị Mạng",
    "7. Cân Điện Tử",
    "8. Thiết Bị Khác",
]

EXCLUDED_KEYWORDS = ("server", "máy chủ", "nas", "san")

COMPUTER_FORM_EXCLUDED = {
    "Mã Tỉnh",
    "Tên Tỉnh",
    "Số thẻ\n(Số mã trên phiếu kê tài sản)",
    "Tên tài sản\n(Theo danh mục CCDC)",
    "Nước sản xuất",
    "Đơn vị bảo hành\n(Nhà cung cấp)",
    "Năm sử dụng",
    "Thời gian bảo hành (Còn/Hết)",
    "Tình trạng",
}

ATTACHMENT_TYPES = [
    {"id": "laser_printer", "name": "Máy in laser", "category": "4A. Máy In Laser"},
    {"id": "thermal_printer", "name": "Máy in nhiệt", "category": "4B. Máy In Nhiệt"},
    {"id": "photocopier", "name": "Máy photocopy", "category": "4C. Máy Photocopy"},
    {"id": "barcode_reader", "name": "Đầu đọc mã vạch", "category": "5. Đầu Đọc Mã Vạch"},
    {"id": "scanner", "name": "Máy quét (Scanner)", "category": "8. Thiết Bị Khác"},
    {"id": "ups", "name": "Bộ lưu điện (UPS)", "category": "8. Thiết Bị Khác"},
]


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def ensure_dirs() -> None:
    for path in (DATA_DIR, UPLOAD_DIR, EXPORT_DIR, CLIENT_INFO_DIR):
        path.mkdir(parents=True, exist_ok=True)
    if not STORE_PATH.exists():
        save_store({"records": []})


def json_bytes(value: object) -> bytes:
    return json.dumps(value, ensure_ascii=False, indent=None).encode("utf-8")


def admin_session_token(expires_at: int) -> str:
    payload = str(expires_at)
    signature = hmac.new(ADMIN_SESSION_SECRET, payload.encode("ascii"), hashlib.sha256).hexdigest()
    return f"{payload}.{signature}"


def valid_admin_session(value: str | None) -> bool:
    if not value:
        return False
    try:
        expires_at, signature = value.split(".", 1)
        if int(expires_at) < int(time.time()):
            return False
    except (TypeError, ValueError):
        return False
    expected = hmac.new(ADMIN_SESSION_SECRET, expires_at.encode("ascii"), hashlib.sha256).hexdigest()
    return hmac.compare_digest(signature, expected)


def load_store() -> dict:
    ensure_dirs()
    try:
        return json.loads(STORE_PATH.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        backup = STORE_PATH.with_suffix(f".broken-{int(time.time())}.json")
        if STORE_PATH.exists():
            STORE_PATH.replace(backup)
        return {"records": []}


def save_store(payload: dict) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    tmp = STORE_PATH.with_suffix(".tmp")
    tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(STORE_PATH)


def clean_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def normalize_client_ip(value: object) -> str:
    address = clean_value(value)
    if address.startswith("::ffff:"):
        return address[7:]
    return address.split("%", 1)[0]


def load_client_inventory_store() -> dict:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with CLIENT_INVENTORY_LOCK:
        if not CLIENT_INVENTORY_PATH.exists():
            return {"clients": {}}
        try:
            payload = json.loads(CLIENT_INVENTORY_PATH.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return {"clients": {}}
        if not isinstance(payload, dict) or not isinstance(payload.get("clients"), dict):
            return {"clients": {}}
        return payload


def save_managed_client_inventory(client_ip: str, fields: dict, machine_id: object = "") -> dict:
    client_ip = normalize_client_ip(client_ip)
    now = int(time.time())
    record = {
        "clientIp": client_ip,
        "machineId": clean_value(machine_id),
        "reportedAt": now_iso(),
        "reportedAtEpoch": now,
        "fields": {str(key): clean_value(value) for key, value in fields.items()},
    }
    with CLIENT_INVENTORY_LOCK:
        store = load_client_inventory_store()
        store["clients"][client_ip] = record
        tmp = CLIENT_INVENTORY_PATH.with_suffix(".tmp")
        tmp.write_text(json.dumps(store, ensure_ascii=False, indent=2), encoding="utf-8")
        tmp.replace(CLIENT_INVENTORY_PATH)
    return record


def load_managed_client_inventory(client_ip: str) -> dict | None:
    client_ip = normalize_client_ip(client_ip)
    store = load_client_inventory_store()
    record = store.get("clients", {}).get(client_ip)
    if not isinstance(record, dict):
        return None
    reported_at = int(record.get("reportedAtEpoch", 0) or 0)
    if reported_at < int(time.time()) - CLIENT_INVENTORY_TTL_SECONDS:
        return None
    return record


def safe_client_token(value: str) -> str | None:
    token = str(value or "").strip().lower()
    if len(token) != 32 or any(char not in "0123456789abcdef" for char in token):
        return None
    return token


def client_info_path(token: str) -> Path:
    return CLIENT_INFO_DIR / f"{token}.json"


def clean_expired_client_info() -> None:
    cutoff = time.time() - CLIENT_INFO_TTL_SECONDS
    for path in CLIENT_INFO_DIR.glob("*.json"):
        try:
            if path.stat().st_mtime < cutoff:
                path.unlink()
        except OSError:
            continue


def save_client_info(token: str, fields: dict) -> None:
    clean_expired_client_info()
    payload = {
        "token": token,
        "createdAt": now_iso(),
        "fields": {str(key): clean_value(value) for key, value in fields.items()},
    }
    target = client_info_path(token)
    tmp = target.with_suffix(".tmp")
    tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(target)


def load_client_info(token: str, consume: bool = True) -> dict | None:
    clean_expired_client_info()
    path = client_info_path(token)
    if not path.exists():
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    if consume:
        path.unlink(missing_ok=True)
    return payload


def parse_multipart_upload(content_type: str, body: bytes, field_name: str = "file") -> tuple[str, bytes] | None:
    if not content_type.lower().startswith("multipart/form-data"):
        return None
    message = BytesParser(policy=policy.default).parsebytes(
        b"Content-Type: "
        + content_type.encode("latin-1")
        + b"\r\nMIME-Version: 1.0\r\n\r\n"
        + body
    )
    if not message.is_multipart():
        return None
    for part in message.iter_parts():
        if part.get_content_disposition() != "form-data":
            continue
        if part.get_param("name", header="content-disposition") != field_name:
            continue
        filename = part.get_filename()
        if not filename:
            return None
        return Path(filename).name, part.get_payload(decode=True) or b""
    return None


def normalize_header(value: object) -> str:
    return clean_value(value).replace("\r\n", "\n")


def visible_headers(ws) -> list[str]:
    headers = []
    for column in range(1, ws.max_column + 1):
        header = normalize_header(ws.cell(4, column).value)
        if header:
            headers.append(header)
    return headers


@lru_cache(maxsize=1)
def template_schema() -> list[dict]:
    wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=False)
    schema = []
    for sheet_name in CATEGORY_ORDER:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        fields = [header for header in visible_headers(ws) if header != "STT"]
        form_fields = fields
        if sheet_name == "3. Máy Tính":
            form_fields = [field for field in fields if field not in COMPUTER_FORM_EXCLUDED]
        schema.append(
            {
                "id": sheet_name,
                "name": sheet_name.split(". ", 1)[-1],
                "sheet": sheet_name,
                "fields": fields,
                "formFields": form_fields,
                "required": [field for field in form_fields if field in COMMON_REQUIRED],
            }
        )
    return schema


def category_by_id(schema: list[dict]) -> dict[str, dict]:
    return {item["id"]: item for item in schema}


def find_record(records: list[dict], record_id: str) -> dict | None:
    return next((record for record in records if record.get("id") == record_id), None)


def record_identity(record: dict) -> str:
    fields = record.get("fields", {})
    parts = [
        record.get("category", ""),
        fields.get("Số thẻ\n(Số mã trên phiếu kê tài sản)", ""),
        fields.get("Serial Number", ""),
        fields.get("Tên máy (Hostname)", ""),
        fields.get("Tên Bưu cục", ""),
        fields.get("Tên tài sản\n(Theo danh mục CCDC)", ""),
    ]
    return "|".join(str(part).strip().lower() for part in parts if str(part).strip())


def clean_attachments(value: object) -> list[dict]:
    allowed = {item["id"]: item for item in ATTACHMENT_TYPES}
    output = []
    if not isinstance(value, list):
        return output
    for item in value:
        if not isinstance(item, dict) or item.get("type") not in allowed:
            continue
        definition = allowed[item["type"]]
        output.append(
            {
                "type": definition["id"],
                "label": definition["name"],
                "category": definition["category"],
                "name": clean_value(item.get("name")),
                "serial": clean_value(item.get("serial")),
            }
        )
    return output


def expanded_records(records: list[dict]) -> list[dict]:
    output = list(records)
    copied_fields = [
        "Mã BĐ Xã",
        "Tên Xã",
        "Mã Bưu cục",
        "Tên Bưu cục",
        "Họ và tên người sử dụng",
        "Mã HRM",
        "Bộ phận / Phòng ban",
    ]
    for record in records:
        if record.get("category") != "3. Máy Tính":
            continue
        source = record.get("fields", {})
        for attachment in record.get("attachments", []):
            fields = {field: source.get(field, "") for field in copied_fields}
            fields.update(
                {
                    "Tên tài sản\n(Theo danh mục CCDC)": attachment.get("label", ""),
                    "Model": attachment.get("name", ""),
                    "Serial Number": attachment.get("serial", ""),
                    "Loại thiết bị": attachment.get("label", ""),
                }
            )
            output.append(
                {
                    "id": f"{record.get('id', '')}:{attachment.get('type', '')}",
                    "category": attachment.get("category", "8. Thiết Bị Khác"),
                    "fields": fields,
                    "derived": True,
                }
            )
    return output


def upsert_records(new_records: list[dict]) -> tuple[int, int]:
    store = load_store()
    records = store.setdefault("records", [])
    index = {record_identity(record): record for record in records if record_identity(record)}
    inserted = 0
    updated = 0
    for new_record in new_records:
        key = record_identity(new_record)
        existing = index.get(key) if key else None
        if existing:
            existing["fields"].update(new_record.get("fields", {}))
            existing["category"] = new_record["category"]
            existing["updatedAt"] = now_iso()
            updated += 1
        else:
            new_record.setdefault("id", uuid4().hex)
            new_record.setdefault("createdAt", now_iso())
            new_record["updatedAt"] = now_iso()
            records.append(new_record)
            if key:
                index[key] = new_record
            inserted += 1
    save_store(store)
    return inserted, updated


def parse_template_excel(path: Path) -> list[dict]:
    schema = template_schema()
    allowed = category_by_id(schema)
    wb = openpyxl.load_workbook(path, data_only=True)
    output: list[dict] = []
    for sheet_name in CATEGORY_ORDER:
        if sheet_name not in wb.sheetnames or sheet_name not in allowed:
            continue
        ws = wb[sheet_name]
        headers = visible_headers(ws)
        if not headers:
            continue
        for row in range(5, ws.max_row + 1):
            stt_value = clean_value(ws.cell(row, 1).value).lower()
            if stt_value.startswith("ví dụ") or stt_value.startswith("vi du"):
                continue
            values = {
                headers[column - 1]: clean_value(ws.cell(row, column).value)
                for column in range(1, min(len(headers), ws.max_column) + 1)
                if headers[column - 1] != "STT"
            }
            if not any(values.values()):
                continue
            if not any(values.get(field, "") for field in ("Tên tài sản\n(Theo danh mục CCDC)", "Serial Number", "Tên máy (Hostname)", "Tên Bưu cục")):
                continue
            output.append({"category": sheet_name, "fields": values})
    return output


def copy_row_style(ws, source_row: int, target_row: int, max_column: int) -> None:
    for column in range(1, max_column + 1):
        src = ws.cell(source_row, column)
        dst = ws.cell(target_row, column)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)


def export_excel(records: list[dict]) -> Path:
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    grouped: dict[str, list[dict]] = {name: [] for name in CATEGORY_ORDER}
    for record in expanded_records(records):
        grouped.setdefault(record.get("category", ""), []).append(record)

    for sheet_name in CATEGORY_ORDER:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = visible_headers(ws)
        max_column = len(headers)
        current_rows = max(ws.max_row - 4, 0)
        needed_rows = max(len(grouped.get(sheet_name, [])), 1)
        if needed_rows > current_rows:
            ws.insert_rows(ws.max_row + 1, amount=needed_rows - current_rows)
        for row in range(5, ws.max_row + 1):
            for column in range(1, max_column + 1):
                ws.cell(row, column).value = None
            copy_row_style(ws, 5, row, max_column)
        for index, record in enumerate(grouped.get(sheet_name, []), start=1):
            row = 4 + index
            fields = record.get("fields", {})
            for column, header in enumerate(headers, start=1):
                ws.cell(row, column).value = index if header == "STT" else fields.get(header, "")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output = EXPORT_DIR / f"Ra_Soat_Thiet_Bi_CNTT_Tong_Hop_{timestamp}.xlsx"
    wb.save(output)
    return output


def run_powershell(script: str) -> str:
    command = ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script]
    try:
        proc = subprocess.run(command, capture_output=True, text=True, timeout=12, encoding="utf-8", errors="replace")
    except (OSError, subprocess.TimeoutExpired):
        return ""
    if proc.returncode != 0:
        return ""
    return proc.stdout.strip()


def local_computer_info() -> dict:
    info = {
        "Tên máy (Hostname)": socket.gethostname(),
        "Hệ điều hành": platform.platform(),
    }
    script = r"""
    $cpu = Get-CimInstance Win32_Processor | Select-Object -First 1 -ExpandProperty Name
    $cs = Get-CimInstance Win32_ComputerSystem | Select-Object -First 1
    $bios = Get-CimInstance Win32_BIOS | Select-Object -First 1
    $os = Get-CimInstance Win32_OperatingSystem | Select-Object -First 1
    $disk = Get-CimInstance Win32_LogicalDisk -Filter "DriveType=3" | Measure-Object -Property Size -Sum
    $adapter = Get-CimInstance Win32_NetworkAdapterConfiguration | Where-Object { $_.IPEnabled -eq $true -and $_.IPAddress } | Select-Object -First 1
    [PSCustomObject]@{
      Cpu=$cpu
      Manufacturer=$cs.Manufacturer
      Model=$cs.Model
      RamGb=[math]::Round($cs.TotalPhysicalMemory / 1GB, 1)
      Serial=$bios.SerialNumber
      Os=$os.Caption
      OsVersion=$os.Version
      DiskGb=[math]::Round($disk.Sum / 1GB, 0)
      Ip=($adapter.IPAddress | Where-Object { $_ -match '^\d+\.' } | Select-Object -First 1)
      Mac=$adapter.MACAddress
    } | ConvertTo-Json -Compress
    """
    raw = run_powershell(script)
    try:
        payload = json.loads(raw) if raw else {}
    except json.JSONDecodeError:
        payload = {}
    if payload:
        info.update(
            {
                "CPU": clean_value(payload.get("Cpu")),
                "Hãng": clean_value(payload.get("Manufacturer")),
                "Model": clean_value(payload.get("Model")),
                "RAM": f"{payload.get('RamGb')} GB" if payload.get("RamGb") else "",
                "Ổ cứng": f"{payload.get('DiskGb')} GB" if payload.get("DiskGb") else "",
                "Serial Number": clean_value(payload.get("Serial")),
                "Hệ điều hành": " ".join(part for part in [clean_value(payload.get("Os")), clean_value(payload.get("OsVersion"))] if part),
                "IP Address": clean_value(payload.get("Ip")),
                "MAC Address": clean_value(payload.get("Mac")),
            }
        )
    return info


def compute_stats(records: list[dict]) -> dict:
    expanded = expanded_records(records)
    by_category = {name: 0 for name in CATEGORY_ORDER}
    by_status: dict[str, int] = {}
    by_unit: dict[str, int] = {}
    missing_required = 0
    excluded_candidates = 0
    monitoring_yes = 0
    monitoring_total = 0
    due_calibration = 0

    schema_by_id = category_by_id(template_schema())
    for record in expanded:
        category = record.get("category", "")
        fields = record.get("fields", {})
        by_category[category] = by_category.get(category, 0) + 1
        status = fields.get("Tình trạng") or fields.get("Tình trạng hoạt động") or "Chưa nhập tình trạng"
        by_status[status] = by_status.get(status, 0) + 1
        unit = fields.get("Tên Bưu cục") or fields.get("Bộ phận / Phòng ban") or "Chưa xác định"
        by_unit[unit] = by_unit.get(unit, 0) + 1
        schema = schema_by_id.get(category)
        if not record.get("derived") and schema and any(not fields.get(field) for field in schema["required"]):
            missing_required += 1
        searchable = " ".join(str(value).lower() for value in fields.values())
        if any(keyword in searchable for keyword in EXCLUDED_KEYWORDS):
            excluded_candidates += 1
        monitor = fields.get("Trạng thái giám sát (Zabbix/ManageEngine...)")
        if monitor:
            monitoring_total += 1
            if "có" in monitor.lower() or "zabbix" in monitor.lower() or "manage" in monitor.lower():
                monitoring_yes += 1
        next_check = fields.get("Thời hạn kiểm định tiếp theo", "")
        if category == "7. Cân Điện Tử" and next_check:
            try:
                parsed = datetime.fromisoformat(next_check[:10])
                if (parsed - datetime.now()).days <= 30:
                    due_calibration += 1
            except ValueError:
                pass

    return {
        "total": len(expanded),
        "surveyForms": len(records),
        "attachedDevices": len(expanded) - len(records),
        "byCategory": by_category,
        "byStatus": by_status,
        "byUnit": dict(sorted(by_unit.items(), key=lambda item: item[1], reverse=True)[:12]),
        "missingRequired": missing_required,
        "excludedCandidates": excluded_candidates,
        "monitoringTotal": monitoring_total,
        "monitoringYes": monitoring_yes,
        "dueCalibration": due_calibration,
    }


class InventoryHandler(BaseHTTPRequestHandler):
    server_version = "DeviceInventoryTool/1.0"

    def log_message(self, fmt: str, *args) -> None:
        print(f"[{self.log_date_time_string()}] {fmt % args}")

    def send_json(
        self,
        payload: object,
        status: HTTPStatus = HTTPStatus.OK,
        headers: dict[str, str] | None = None,
    ) -> None:
        body = json_bytes(payload)
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Cache-Control", "no-store")
        if headers:
            for name, value in headers.items():
                self.send_header(name, value)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def is_admin(self) -> bool:
        cookie = SimpleCookie()
        cookie.load(self.headers.get("Cookie", ""))
        morsel = cookie.get(ADMIN_COOKIE_NAME)
        return valid_admin_session(morsel.value if morsel else None)

    def require_admin(self) -> bool:
        if self.is_admin():
            return True
        self.send_json({"error": "Yêu cầu đăng nhập quản trị."}, HTTPStatus.FORBIDDEN)
        return False

    def send_file(self, path: Path, download_name: str | None = None) -> None:
        if not path.exists() or not path.is_file():
            self.send_json({"error": "Không tìm thấy file."}, HTTPStatus.NOT_FOUND)
            return
        body = path.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", mimetypes.guess_type(path.name)[0] or "application/octet-stream")
        self.send_header("Content-Length", str(len(body)))
        if download_name:
            self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{download_name}")
        self.end_headers()
        self.wfile.write(body)

    def send_static(self, route: str) -> None:
        route = "index.html" if route in ("", "/") else unquote(route).lstrip("/")
        path = (STATIC_DIR / route).resolve()
        if STATIC_DIR.resolve() not in path.parents and path != STATIC_DIR.resolve():
            self.send_error(HTTPStatus.FORBIDDEN)
            return
        self.send_file(path)

    def read_json_body(self) -> dict:
        length = int(self.headers.get("Content-Length", "0") or 0)
        if length <= 0:
            return {}
        body = self.rfile.read(length)
        return json.loads(body.decode("utf-8"))

    def client_ip(self) -> str:
        return normalize_client_ip(self.client_address[0])

    def do_GET(self) -> None:  # noqa: N802
        parsed = urlparse(self.path)
        if parsed.path == "/api/health":
            self.send_json(
                {
                    "status": "ok",
                    "time": now_iso(),
                    "host": self.server.server_address[0],
                    "port": self.server.server_address[1],
                    "managedClientInventory": True,
                }
            )
        elif parsed.path == "/api/metadata":
            self.send_json(
                {
                    "schema": template_schema(),
                    "deadline": "2026-06-15",
                    "attachmentTypes": ATTACHMENT_TYPES,
                }
            )
        elif parsed.path == "/api/admin/session":
            self.send_json({"authenticated": self.is_admin()})
        elif parsed.path == "/api/records":
            if not self.require_admin():
                return
            store = load_store()
            self.send_json({"records": store.get("records", []), "stats": compute_stats(store.get("records", []))})
        elif parsed.path == "/api/stats":
            if not self.require_admin():
                return
            self.send_json(compute_stats(load_store().get("records", [])))
        elif parsed.path == "/api/computer-info":
            self.send_json(
                {
                    "error": "Không lấy cấu hình qua máy chủ. Hãy chạy Bộ thu thập thông tin trên máy client.",
                },
                HTTPStatus.GONE,
            )
        elif parsed.path == "/api/client-inventory/current":
            payload = load_managed_client_inventory(self.client_ip())
            if payload is None:
                self.send_json(
                    {
                        "error": "Máy này chưa gửi cấu hình tự động hoặc dữ liệu đã quá 60 phút.",
                        "clientIp": self.client_ip(),
                    },
                    HTTPStatus.NOT_FOUND,
                )
                return
            self.send_json(payload)
        elif parsed.path == "/api/client-info":
            token = safe_client_token(parse_qs(parsed.query).get("id", [""])[0])
            if not token:
                self.send_json({"error": "Mã phiên máy client không hợp lệ."}, HTTPStatus.BAD_REQUEST)
                return
            payload = load_client_info(token)
            if payload is None:
                self.send_json(
                    {"error": "Chưa nhận được thông tin máy client hoặc mã phiên đã hết hạn."},
                    HTTPStatus.NOT_FOUND,
                )
                return
            self.send_json(payload)
        elif parsed.path == "/client-agent.zip":
            collector_paths = [
                APP_ROOT / "client_collector.ps1",
                APP_ROOT / "client_collector.vbs",
                APP_ROOT / "client_collector.bat",
            ]
            if any(not path.exists() for path in collector_paths):
                self.send_json({"error": "Chưa có bộ thu thập máy client."}, HTTPStatus.NOT_FOUND)
                return
            buffer = io.BytesIO()
            with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
                for path in collector_paths:
                    archive.write(path, path.name)
            body = buffer.getvalue()
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/zip")
            self.send_header("Content-Disposition", "attachment; filename=VNPost_Thu_Thap_Thong_Tin_May.zip")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
        elif parsed.path == "/api/export-excel":
            if not self.require_admin():
                return
            output = export_excel(load_store().get("records", []))
            self.send_file(output, output.name)
        elif parsed.path == "/api/export-json":
            if not self.require_admin():
                return
            payload = json_bytes(load_store())
            name = f"du_lieu_ra_soat_thiet_bi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Disposition", f"attachment; filename={name}")
            self.send_header("Content-Length", str(len(payload)))
            self.end_headers()
            self.wfile.write(payload)
        else:
            self.send_static(parsed.path)

    def do_POST(self) -> None:  # noqa: N802
        parsed = urlparse(self.path)
        try:
            if parsed.path == "/api/client-inventory":
                length = int(self.headers.get("Content-Length", "0") or 0)
                if length <= 0 or length > MAX_CLIENT_INFO_BYTES:
                    self.send_json({"error": "Dữ liệu agent không hợp lệ."}, HTTPStatus.BAD_REQUEST)
                    return
                payload = self.read_json_body()
                fields = payload.get("fields")
                if not isinstance(fields, dict) or not fields:
                    self.send_json({"error": "Thiếu dữ liệu cấu hình máy."}, HTTPStatus.BAD_REQUEST)
                    return
                record = save_managed_client_inventory(
                    self.client_ip(),
                    fields,
                    payload.get("machineId"),
                )
                self.send_json(
                    {
                        "ok": True,
                        "clientIp": record["clientIp"],
                        "reportedAt": record["reportedAt"],
                    }
                )
            elif parsed.path == "/api/client-info":
                length = int(self.headers.get("Content-Length", "0") or 0)
                if length <= 0 or length > MAX_CLIENT_INFO_BYTES:
                    self.send_json({"error": "Dữ liệu máy client không hợp lệ."}, HTTPStatus.BAD_REQUEST)
                    return
                payload = self.read_json_body()
                token = safe_client_token(payload.get("token"))
                fields = payload.get("fields")
                if not token or not isinstance(fields, dict):
                    self.send_json({"error": "Thiếu mã phiên hoặc dữ liệu cấu hình."}, HTTPStatus.BAD_REQUEST)
                    return
                save_client_info(token, fields)
                self.send_json({"ok": True, "token": token})
            elif parsed.path == "/api/admin/login":
                payload = self.read_json_body()
                if not hmac.compare_digest(str(payload.get("password", "")), ADMIN_PASSWORD):
                    self.send_json({"error": "Mật khẩu quản trị không đúng."}, HTTPStatus.UNAUTHORIZED)
                    return
                expires_at = int(time.time()) + ADMIN_SESSION_SECONDS
                token = admin_session_token(expires_at)
                cookie = f"{ADMIN_COOKIE_NAME}={token}; Path=/; HttpOnly; SameSite=Strict; Max-Age={ADMIN_SESSION_SECONDS}"
                self.send_json({"authenticated": True}, headers={"Set-Cookie": cookie})
            elif parsed.path == "/api/admin/logout":
                cookie = f"{ADMIN_COOKIE_NAME}=; Path=/; HttpOnly; SameSite=Strict; Max-Age=0"
                self.send_json({"authenticated": False}, headers={"Set-Cookie": cookie})
            elif parsed.path == "/api/records":
                payload = self.read_json_body()
                schema = category_by_id(template_schema())
                category = payload.get("category")
                if category not in schema:
                    self.send_json({"error": "Nhóm thiết bị không hợp lệ."}, HTTPStatus.BAD_REQUEST)
                    return
                if payload.get("id") and not self.require_admin():
                    return
                record = {
                    "id": payload.get("id") or uuid4().hex,
                    "category": category,
                    "fields": {field: clean_value(value) for field, value in payload.get("fields", {}).items()},
                    "attachments": clean_attachments(payload.get("attachments")) if category == "3. Máy Tính" else [],
                    "createdAt": payload.get("createdAt") or now_iso(),
                    "updatedAt": now_iso(),
                }
                store = load_store()
                records = store.setdefault("records", [])
                existing = find_record(records, record["id"])
                if existing:
                    existing.update(record)
                else:
                    records.append(record)
                save_store(store)
                response = {"ok": True, "recordId": record["id"]}
                if self.is_admin():
                    response.update({"record": record, "stats": compute_stats(records)})
                self.send_json(response)
            elif parsed.path == "/api/import-excel":
                if not self.require_admin():
                    return
                self.handle_import()
            elif parsed.path == "/api/reset":
                if not self.require_admin():
                    return
                save_store({"records": []})
                self.send_json({"ok": True})
            else:
                self.send_json({"error": "API không hỗ trợ."}, HTTPStatus.NOT_FOUND)
        except Exception as exc:  # pragma: no cover - API safety
            self.send_json({"error": str(exc)}, HTTPStatus.INTERNAL_SERVER_ERROR)

    def do_DELETE(self) -> None:  # noqa: N802
        parsed = urlparse(self.path)
        if parsed.path != "/api/records":
            self.send_json({"error": "API không hỗ trợ."}, HTTPStatus.NOT_FOUND)
            return
        if not self.require_admin():
            return
        record_id = parse_qs(parsed.query).get("id", [""])[0]
        store = load_store()
        before = len(store.get("records", []))
        store["records"] = [record for record in store.get("records", []) if record.get("id") != record_id]
        save_store(store)
        self.send_json({"deleted": before - len(store["records"]), "stats": compute_stats(store["records"])})

    def handle_import(self) -> None:
        length = int(self.headers.get("Content-Length", "0") or 0)
        if length <= 0 or length > MAX_UPLOAD_BYTES:
            self.send_json({"error": "File rỗng hoặc vượt quá 50MB."}, HTTPStatus.BAD_REQUEST)
            return
        body = self.rfile.read(length)
        uploaded = parse_multipart_upload(self.headers.get("Content-Type", ""), body)
        if uploaded is None:
            self.send_json({"error": "Chưa nhận được file Excel."}, HTTPStatus.BAD_REQUEST)
            return
        filename, file_bytes = uploaded
        if not filename.lower().endswith(".xlsx"):
            self.send_json({"error": "Chỉ nhận file định dạng .xlsx."}, HTTPStatus.BAD_REQUEST)
            return
        target = UPLOAD_DIR / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}"
        target.write_bytes(file_bytes)
        parsed = parse_template_excel(target)
        inserted, updated = upsert_records(parsed)
        records = load_store().get("records", [])
        self.send_json({"inserted": inserted, "updated": updated, "read": len(parsed), "stats": compute_stats(records)})


def lan_addresses(port: int) -> list[str]:
    addresses = [f"http://127.0.0.1:{port}/"]
    try:
        for ip in socket.gethostbyname_ex(socket.gethostname())[2]:
            if ip and not ip.startswith("127."):
                addresses.append(f"http://{ip}:{port}/")
    except OSError:
        pass
    return sorted(set(addresses))


def main() -> None:
    if not TEMPLATE_PATH.exists():
        raise SystemExit(f"Khong tim thay file mau: {TEMPLATE_PATH}")
    ensure_dirs()
    port = int(os.environ.get("DEVICE_INVENTORY_PORT") or os.environ.get("PORT") or "8789")
    default_host = "0.0.0.0" if os.environ.get("PORT") else "127.0.0.1"
    host = os.environ.get("DEVICE_INVENTORY_HOST", default_host)
    public_bind = host not in {"127.0.0.1", "localhost", "::1"}
    if public_bind and ADMIN_PASSWORD == "250389" and "DEVICE_INVENTORY_ADMIN_PASSWORD" not in os.environ:
        raise SystemExit(
            "Tu choi chay public voi mat khau mac dinh. "
            "Hay dat bien moi truong DEVICE_INVENTORY_ADMIN_PASSWORD."
        )
    server = ThreadingHTTPServer((host, port), InventoryHandler)
    print("Cong cu ra soat thiet bi CNTT dang chay:", flush=True)
    for address in lan_addresses(port):
        print(f"- {address}", flush=True)
    print("Nhan Ctrl+C de dung.", flush=True)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nDa dung server.")


if __name__ == "__main__":
    main()
